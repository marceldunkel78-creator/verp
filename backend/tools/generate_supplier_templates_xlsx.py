"""
Generiert XLSX-Templates mit Dropdowns für Supplier und SupplierContacts.
Benötigt openpyxl: pip install openpyxl
Ausführen: python tools/generate_supplier_templates_xlsx.py
"""
import os
import sys
from pathlib import Path

# Make sure the backend folder is on sys.path
BACKEND_DIR = Path(__file__).resolve().parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'verp.settings')
import django
django.setup()

from suppliers.models import Supplier, SupplierContact
from verp_settings.models import DeliveryTerm, PaymentTerm

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / 'Datenvorlagen' / 'templates'
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Prepare lists
suppliers = list(Supplier.objects.all().order_by('company_name'))
contact_types = [k for k, _ in SupplierContact.CONTACT_TYPE_CHOICES]
delivery_terms = [d.incoterm for d in DeliveryTerm.objects.all().order_by('incoterm')]
payment_terms = [p.name for p in PaymentTerm.objects.all().order_by('name')]

# Suppliers template
wb = Workbook()
ws = wb.active
ws.title = 'Suppliers'
# Insert a helper column `current_supplier_name` next to `supplier_number` so users immediately see the existing name
headers = ['supplier_number','current_supplier_name','company_name','street','house_number','address_supplement','postal_code','city','state','country','email','phone','website','customer_number','payment_term','delivery_term','delivery_instruction','notes','is_active']
ws.append(headers)

# Add suppliers list as hidden sheet for dropdown + VLOOKUP source
if suppliers:
    s_ws = wb.create_sheet(title='__suppliers_list')
    for i, s in enumerate(suppliers, start=1):
        # Column A contains a combined display value 'number - name' for the dropdown
        s_ws.cell(row=i, column=1, value=f"{s.supplier_number} - {s.company_name}")
        # Column B keeps the supplier name for VLOOKUP
        s_ws.cell(row=i, column=2, value=s.company_name)
    s_ws.sheet_state = 'hidden'

# Create data validation for supplier_number to pick existing suppliers and fill helper names via VLOOKUP
if suppliers:
    dv_sup = DataValidation(type='list', formula1=f"=__suppliers_list!$A$1:$A${len(suppliers)}", allow_blank=True)
    ws.add_data_validation(dv_sup)
    dv_sup.ranges.add(f"{get_column_letter(1)}2:{get_column_letter(1)}1048576")
    # Pre-fill VLOOKUP formulas for helper column (first 1000 rows)
    for row in range(2, 1002):
        ws.cell(row=row, column=2, value=f"=IFERROR(VLOOKUP($A{row},__suppliers_list!$A$1:$B${len(suppliers)},2,FALSE),\"\")")

# Add delivery terms and payment terms hidden lists
dt_ws = wb.create_sheet(title='__delivery_terms')
for i, v in enumerate(delivery_terms, start=1):
    dt_ws.cell(row=i, column=1, value=v)
dt_ws.sheet_state = 'hidden'

pt_ws = wb.create_sheet(title='__payment_terms')
for i, v in enumerate(payment_terms, start=1):
    pt_ws.cell(row=i, column=1, value=v)
pt_ws.sheet_state = 'hidden'

# Create data validation for payment_term and delivery_term
if payment_terms:
    dv_pt = DataValidation(type='list', formula1=f"=__payment_terms!$A$1:$A${len(payment_terms)}", allow_blank=True)
    ws.add_data_validation(dv_pt)
    # apply to column 15 (payment_term) — index shifted by helper column
    dv_pt.ranges.add(f"{get_column_letter(15)}2:{get_column_letter(15)}1048576")

if delivery_terms:
    dv_dt = DataValidation(type='list', formula1=f"=__delivery_terms!$A$1:$A${len(delivery_terms)}", allow_blank=True)
    ws.add_data_validation(dv_dt)
    dv_dt.ranges.add(f"{get_column_letter(16)}2:{get_column_letter(16)}1048576")

# Save suppliers template
wb.save(OUT_DIR / 'suppliers_template.xlsx')

# Supplier Contacts template
wb2 = Workbook()
ws2 = wb2.active
ws2.title = 'Contacts'
headers2 = ['supplier_number','contact_type','is_primary','contact_person','contact_function','street','house_number','address_supplement','postal_code','city','state','country','email','phone','mobile','notes','is_active']
ws2.append(headers2)

# Add contact types list
ct_ws = wb2.create_sheet(title='__contact_types')
for i, v in enumerate(contact_types, start=1):
    ct_ws.cell(row=i, column=1, value=v)
ct_ws.sheet_state = 'hidden'

dv_ct = DataValidation(type='list', formula1=f"=__contact_types!$A$1:$A${len(contact_types)}", allow_blank=False)
ws2.add_data_validation(dv_ct)
dv_ct.ranges.add(f"{get_column_letter(2)}2:{get_column_letter(2)}1048576")

# Supplier dropdown (if suppliers exist)
if suppliers:
    # create a hidden sheet with combined 'number - name' values for dropdown
    s2_ws = wb2.create_sheet(title='__suppliers_list')
    for i, s in enumerate(suppliers, start=1):
        s2_ws.cell(row=i, column=1, value=f"{s.supplier_number} - {s.company_name}")
    s2_ws.sheet_state = 'hidden'
    dv_sup = DataValidation(type='list', formula1=f"=__suppliers_list!$A$1:$A${len(suppliers)}", allow_blank=False)
    ws2.add_data_validation(dv_sup)
    dv_sup.ranges.add(f"{get_column_letter(1)}2:{get_column_letter(1)}1048576")

# delivery instruction left free (text) - not validated

wb2.save(OUT_DIR / 'supplier_contacts_template.xlsx')

print('XLSX templates generated in', OUT_DIR)
print(' - suppliers_template.xlsx')
print(' - supplier_contacts_template.xlsx')
print('\nNote: Install openpyxl if missing: pip install openpyxl')